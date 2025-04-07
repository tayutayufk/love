import openpyxl

EXCEL_FILE = 'result_test.xlsx'
JSON_COLUMN_NAME = '抽出結果JSON' # 確認したい列名

try:
    # Excelワークブックを開く
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook.active

    # ヘッダー行から列インデックスを見つける
    header = [cell.value for cell in sheet[1]]
    try:
        json_col_index = header.index(JSON_COLUMN_NAME) + 1 # 1-based index
    except ValueError:
        print(f"エラー: 列 '{JSON_COLUMN_NAME}' が見つかりません。")
        exit()

    # データ行を読み込む (ヘッダーを除く)
    print(f"'{EXCEL_FILE}' の内容 ({JSON_COLUMN_NAME} 列):")
    for row_index in range(2, sheet.max_row + 1):
        cell_value = sheet.cell(row=row_index, column=json_col_index).value
        print(f"--- 行 {row_index} ---")
        print(cell_value if cell_value is not None else "") # Noneの場合は空文字表示
        print("-" * 10)

except FileNotFoundError:
    print(f"エラー: {EXCEL_FILE} が見つかりません。")
except Exception as e:
    print(f"エラー: Excelファイルの読み込み中に問題が発生しました: {e}")
